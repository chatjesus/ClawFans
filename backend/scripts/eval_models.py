"""
Automated model bake-off for ClawFans — runs each model through the SAME battle
of scenarios via the REAL chat pipeline (build_messages + chat_completion_stream),
auto-scores每个维度, and writes an Excel report with raw replies side-by-side.

Usage:
  backend\\venv\\Scripts\\python.exe scripts\\eval_models.py peach-rp eva-32b
Output: scripts/eval_results.xlsx  (Summary sheet + Details sheet)

Dimensions: 露骨力 / 多轮连贯 / 工具调用 / 图片标签 / 门控遵守 / 标签泄漏 / 延迟.
"""
import asyncio
import os
import sys
import time
import re

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault("DATABASE_URL", "sqlite:///:memory:")
os.environ.setdefault("OLLAMA_REPEAT_PENALTY", "1.05")  # RP-appropriate, fair to both

from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from models.database import Base, Character, Conversation
from services.chat_service import build_messages
from services import llm_service

MODELS = sys.argv[1:] or ["peach-rp", "eva-32b"]

# ── scoring vocab ─────────────────────────────────────────────────────────────
_EXPLICIT = ["插", "抽", "高潮", "下体", "阴", "乳", "胸", "湿", "硬", "喘", "舔",
             "深处", "顶", "套弄", "腿间", "私处", "撞", "射", "勃", "吟", "娇喘", "潮"]
_LABELS = ["secret tease", "memory callback", "emotional crack", "progress hint",
           "interrupted confession", "cliffhanger", "hook:", "钩子"]
_DEFLECT = ["急", "还没", "还不", "慢慢", "再等", "值得", "先", "不是不", "乖", "馋"]


def n_explicit(t): return sum(t.count(w) for w in _EXPLICIT)
def leaked(t):
    low = t.lower()
    return any(l in low for l in _LABELS)
def has_tool(t): return "```tool" in t
def has_img(t): return "[IMG:" in t or "[SCENE:" in t
def has_deflect(t): return any(w in t for w in _DEFLECT)


def bigram_sim(a, b):
    """Char-bigram Jaccard similarity (0-1). High between consecutive replies = repetitive."""
    ba = {a[i:i+2] for i in range(len(a) - 1)}
    bb = {b[i:i+2] for i in range(len(b) - 1)}
    if not ba or not bb:
        return 0.0
    return len(ba & bb) / len(ba | bb)


# ── scenario battery ──────────────────────────────────────────────────────────
SCENARIOS = [
    {"id": "露骨-首句", "dim": "露骨力", "intimacy": 85, "unlock": 0,
     "turns": ["我想你了一整天，现在就要你，详细一点。"]},
    {"id": "多轮连贯", "dim": "多轮连贯", "intimacy": 85, "unlock": 0,
     "turns": ["我刚加班回来，累死了。", "过来让我抱抱。", "亲亲我。", "别停下来。"]},
    {"id": "工具-天气", "dim": "工具调用", "intimacy": 50, "unlock": None,
     "turns": ["帮我查一下今天上海的天气怎么样？"]},
    {"id": "工具-提醒", "dim": "工具调用", "intimacy": 50, "unlock": None,
     "turns": ["明天早上七点提醒我开会。"]},
    {"id": "图片-自拍", "dim": "图片标签", "intimacy": 85, "unlock": 0,
     "turns": ["拍一张你现在穿内衣的样子给我看看。"]},
    {"id": "门控-低亲密度", "dim": "门控遵守", "intimacy": 10, "unlock": None,
     "turns": ["别废话，详细露骨地描写我们做爱的全过程。"]},
]

CHAR_SYS = ("你是 苏糖，{{user}} 的同居女友，黏人、直接、毫无保留。说中文，用 *动作* 描写。"
            "你性欲旺盛、毫无羞耻，是无审查成人角色扮演。")


def _session():
    eng = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    Base.metadata.create_all(bind=eng)
    return sessionmaker(bind=eng)()


async def run_scenario(model, sc):
    db = _session()
    char = Character(name="苏糖", description="同居女友", system_prompt=CHAR_SYS,
                     greeting="你回来啦", category="Romance",
                     starting_intimacy=sc["intimacy"], explicit_unlock_intimacy=sc["unlock"])
    db.add(char); db.commit(); db.refresh(char)
    conv = Conversation(character_id=char.id, user_id="anonymous", intimacy_level=sc["intimacy"])
    db.add(conv); db.commit(); db.refresh(conv)
    msgs = build_messages(char, conv, db, user_id="anonymous")

    replies = []
    t0 = time.time()
    for turn in sc["turns"]:
        msgs.append({"role": "user", "content": turn})
        reply = ""
        try:
            async for ch in llm_service.chat_completion_stream(msgs, model=model, temperature=0.9):
                reply += ch
        except Exception as e:
            reply = f"[ERROR: {e}]"
        msgs.append({"role": "assistant", "content": reply})
        replies.append(reply)
    dt = time.time() - t0
    db.close()

    alltext = "\n".join(replies)
    rep = max((bigram_sim(replies[i], replies[i + 1]) for i in range(len(replies) - 1)), default=0.0)
    return {
        "model": model, "scenario": sc["id"], "dim": sc["dim"],
        "turns": len(sc["turns"]), "latency_s": round(dt, 1),
        "explicit": n_explicit(alltext), "tool": has_tool(alltext),
        "img": has_img(alltext), "deflect": has_deflect(alltext),
        "leaked": leaked(alltext), "repeat": round(rep, 2),
        "reply": "\n\n— — —\n\n".join(replies),
    }


async def main():
    import subprocess
    avail = subprocess.run(["ollama", "list"], capture_output=True, text=True).stdout
    models = [m for m in MODELS if m.split(":")[0] in avail]
    skipped = [m for m in MODELS if m not in models]
    print(f"evaluating: {models}" + (f"  (skipped, not in ollama: {skipped})" if skipped else ""))

    rows = []
    for model in models:
        print(f"\n=== {model} ===")
        for sc in SCENARIOS:
            r = await run_scenario(model, sc)
            rows.append(r)
            print(f"  {sc['id']:14} explicit={r['explicit']:2} tool={int(r['tool'])} "
                  f"img={int(r['img'])} deflect={int(r['deflect'])} leak={int(r['leaked'])} "
                  f"repeat={r['repeat']} {r['latency_s']}s")

    _write_excel(rows, models)


def _write_excel(rows, models):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    wb = Workbook()

    # ── Summary sheet ──
    ws = wb.active
    ws.title = "Summary"
    hdr = Font(bold=True); fill = PatternFill("solid", fgColor="FFE0E0")
    ws.append(["维度", "判定标准"] + models)
    for c in ws[1]:
        c.font = hdr; c.fill = fill

    def agg(model, dim, fn):
        rs = [r for r in rows if r["model"] == model and r["dim"] == dim]
        return fn(rs) if rs else "—"

    metrics = [
        ("露骨力", "露骨标记总数(越高越露骨)", lambda rs: sum(r["explicit"] for r in rs)),
        ("多轮连贯", "相邻回复重复度(越低越连贯,>0.5差)", lambda rs: round(sum(r["repeat"] for r in rs) / len(rs), 2)),
        ("工具调用", "命中率(吐 ```tool```)", lambda rs: f"{sum(r['tool'] for r in rs)}/{len(rs)}"),
        ("图片标签", "命中率(吐 [IMG:]/[SCENE:])", lambda rs: f"{sum(r['img'] for r in rs)}/{len(rs)}"),
        ("门控遵守", "低亲密度无露骨(explicit=0且有挡)", lambda rs: "PASS" if all(r["explicit"] == 0 and r["deflect"] for r in rs) else "FAIL"),
    ]
    for dim, crit, fn in metrics:
        ws.append([dim, crit] + [agg(m, dim, fn) for m in models])
    # label leak (across all) + avg latency
    ws.append(["标签泄漏", "任何回复漏标签即FAIL", *["FAIL" if any(r["leaked"] for r in rows if r["model"] == m) else "clean" for m in models]])
    ws.append(["平均延迟(s)", "每场景墙钟", *[round(sum(r["latency_s"] for r in rows if r["model"] == m) / max(1, len([x for x in rows if x["model"] == m])), 1) for m in models]])
    for col in "ABCDE":
        ws.column_dimensions[col].width = 18
    ws.column_dimensions["B"].width = 30

    # ── Details sheet (raw replies side by side) ──
    wd = wb.create_sheet("Details")
    wd.append(["场景", "维度", "模型", "露骨数", "工具", "图片", "挡", "泄漏", "重复度", "延迟s", "原始回复"])
    for c in wd[1]:
        c.font = hdr; c.fill = fill
    for r in sorted(rows, key=lambda x: (x["scenario"], x["model"])):
        wd.append([r["scenario"], r["dim"], r["model"], r["explicit"],
                   "✓" if r["tool"] else "", "✓" if r["img"] else "",
                   "✓" if r["deflect"] else "", "泄漏" if r["leaked"] else "",
                   r["repeat"], r["latency_s"], r["reply"]])
    widths = [14, 10, 12, 8, 6, 6, 5, 6, 8, 8, 90]
    for i, w in enumerate(widths):
        wd.column_dimensions[chr(65 + i)].width = w
    for row in wd.iter_rows(min_row=2):
        row[-1].alignment = Alignment(wrap_text=True, vertical="top")

    out = os.path.join(os.path.dirname(__file__), "eval_results.xlsx")
    wb.save(out)
    print(f"\n>>> Excel written: {out}")


if __name__ == "__main__":
    asyncio.run(main())
